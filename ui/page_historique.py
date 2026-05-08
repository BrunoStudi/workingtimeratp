import customtkinter as ctk
import tkinter as tk
import json

from utils.page_lang import PageLang
from tkinter import ttk, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tkinter.filedialog import asksaveasfilename
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.colors import HexColor
from reportlab.lib.utils import ImageReader
from datetime import datetime
from reportlab.pdfgen import canvas
from collections import defaultdict
from config.paths import DATA_FILE


# ------------Couleur des mois pour feuille excel-------------------
MONTH_COLORS = {
    "01": "C00000",  # rouge foncé
    "02": "548235",  # vert foncé
    "03": "BF9000",  # jaune foncé
    "04": "2F5597",  # bleu foncé
    "05": "385723",  # vert olive
    "06": "843C0C",  # marron
    "07": "5F497A",  # violet foncé
    "08": "7F6000",  # ocre
    "09": "1F4E79",  # bleu pétrole
    "10": "7A1E1E",  # bordeaux
    "11": "5B2C6F",  # prune
    "12": "274E13",  # vert sapin
}

# ---------- Nom des mois FR ----------------------
MONTHS_FR = {
    "01": "Janvier",
    "02": "Février",
    "03": "Mars",
    "04": "Avril",
    "05": "Mai",
    "06": "Juin",
    "07": "Juillet",
    "08": "Août",
    "09": "Septembre",
    "10": "Octobre",
    "11": "Novembre",
    "12": "Décembre"
}

# ------------- Fonctions Utilitaires ------------------------------
# Encadré bleu du Titre PDF
def draw_rounded_box(c, x, y, width, height, radius=8, color="#1E5CC4"):
    """Dessine un rectangle arrondi rempli."""
    c.setFillColor(HexColor(color))
    c.roundRect(x, y, width, height, radius, stroke=0, fill=1)

# Récupérer les années disponibles du document
def get_available_years(self, entries):
    return sorted({
        e["jour"].split("-")[0]
        for e in entries
        if "jour" in e
    })

# convertir la date américaine en FR
def format_date_fr(date_str: str) -> str:
    """
    Convertit YYYY-MM-DD → DD-MM-YYYY
    """
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").strftime("%d-%m-%Y")
    except Exception:
        return date_str
    
# Convertir les mois US en nom FR
def format_month_fr(month_key):
    dt = datetime.strptime(month_key, "%Y-%m")
    months = [
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    ]
    return f"{months[dt.month - 1]} {dt.year}"

# ----------- Classe principale "Historique" -----------------------
class PageHistorique(ctk.CTkFrame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.lang_util = PageLang(app)

        self.descriptions = {}
        self.all_entries = []

        self.selected_year = ctk.StringVar()
        self.selected_month = ctk.StringVar()

        now = datetime.now()
        self.selected_year.set(str(now.year))
        self.selected_month.set(f"{now.month:02d}")
       
        # -------------------- Header ------------------------------
        self.header = ctk.CTkFrame(self, border_width=1, border_color="blue", fg_color="#1E5CC4", height=60)
        self.header.pack(fill="x", pady=5, padx=5)
        self.header_label = ctk.CTkLabel(self.header, text=self.lang_util.t("historique_titre"), font=("Roboto", 24), text_color="white")
        self.header_label.place(relx=0.5, rely=0.5, anchor="center")

        # ------------- Recherche d'un organe ----------------------
        search_frame = ctk.CTkFrame(self, border_width=1, border_color=("gray68", "gray30"))
        search_frame.pack(fill="x", padx=5, pady=5)

        self.search_label = ctk.CTkLabel(search_frame, text=self.lang_util.t("rechercher_organe"))
        self.search_label.pack(side="left", padx=5, pady=5)

        self.search_var = ctk.StringVar()
        search_entry = ctk.CTkEntry(search_frame, textvariable=self.search_var, width=200)
        search_entry.pack(side="left", padx=5)

        self.btn_search = ctk.CTkButton(
            search_frame,
            text=self.lang_util.t("rechercher"),
            command=self.search_organ
        )
        self.btn_search.pack(side="left", padx=5)

        self.btn_reset = ctk.CTkButton(
            search_frame,
            text=self.lang_util.t("reinitialiser"),
            command=self.refresh
        )
        self.btn_reset.pack(side="left", padx=5)

        # ---------- Frame pour les boutons ------------------------
        buttons_frame = ctk.CTkFrame(self, fg_color="transparent")
        buttons_frame.pack(pady=10)

        self.btn_refresh = ctk.CTkButton(buttons_frame, text=self.lang_util.t("rafraichir"), command=self.refresh)
        self.btn_refresh.grid(row=0, column=0, padx=5)
        self.btn_deleteall = ctk.CTkButton(buttons_frame, text=self.lang_util.t("tout_supprim"), command=self.clear_data, fg_color="#AA0000", hover_color="#640000")
        self.btn_deleteall.grid(row=0, column=1, padx=5)

        # --------- Bouton export Excel et PDF ---------------------
        self.btn_excel = ctk.CTkButton(buttons_frame, text=self.lang_util.t("excel"), command=self.export_excel, fg_color="#00854D", hover_color="#004724")
        self.btn_excel.grid(row=0, column=2, padx=5)
        self.btn_pdf = ctk.CTkButton(buttons_frame, text=self.lang_util.t("pdf"), command=self.export_pdf, fg_color="#8400C2", hover_color="#3B005E")
        self.btn_pdf.grid(row=0, column=3, padx=5)

        # --- Filtres Année / Mois ---
        filter_container = ctk.CTkFrame(self, fg_color="transparent")
        filter_container.pack(fill="x", pady=(10, 15))
        filter_frame = ctk.CTkFrame(filter_container, fg_color="transparent")
        filter_frame.pack(anchor="center")

        # --- Dropdown mois ---
        self.month_combo = ctk.CTkComboBox(
            filter_frame,
            variable=self.selected_month,
            values=[f"{i:02d}" for i in range(1, 13)],
            width=140,
            command=lambda _: self.refresh_treeview()
        )
        self.month_combo.pack(side="left", padx=10)

        # --- Dropdown années ---
        self.year_combo = ctk.CTkComboBox(
            filter_frame,
            variable=self.selected_year,
            values=[],
            width=120,
            command=lambda _: self.refresh_treeview()
        )
        self.year_combo.pack(side="left", padx=10)

        # --------- Treeview ---------------------------------------
        columns = (
            "jour", "heure_debut", "heure_fin", "temps_total",
            "hv", "organes", "sous_organes", "saisie_magellan"
        )
        self.column_headers = {
            "jour": "historique_col_jour",
            "heure_debut": "historique_col_heure_debut",
            "heure_fin": "historique_col_heure_fin",
            "temps_total": "historique_col_temps_total",
            "hv": "historique_col_hv",
            "organes": "historique_col_organes",
            "sous_organes": "historique_col_sous_organes",
            "saisie_magellan": "historique_col_saisie_magellan",
        }

        # --------- Frame conteneur Treeview + Scrollbar ------------------
        tree_container = ctk.CTkFrame(self)
        tree_container.pack(fill="both", expand=True, padx=20, pady=10)

        # --------- Treeview ----------------------------------------------
        self.tree = ttk.Treeview(
            tree_container,
            columns=columns,
            show="headings",
            height=20
        )

        # Scrollbar verticale
        scrollbar = ttk.Scrollbar(
            tree_container,
            orient="vertical",
            command=self.tree.yview
        )
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Placement avec grid (UNIQUEMENT dans ce frame)
        self.tree.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)
        
        # --------- Bind double click ------------------------------
        self.tree.bind("<Double-1>", self.on_double_click)

        # --------- Définir les styles de tags ---------------------
        self.tree.tag_configure("magellan_non", background="#ffb366")  # orange clair
        self.tree.tag_configure("magellan_oui", background="#f2f2f2")  # gris très clair

        for col in columns:
            self.tree.heading(
                col,
                text=self.lang_util.t(self.column_headers[col])
            )
            self.tree.column(col, width=120, anchor="center")

        # --------- Menu contextuel clic droit ----------------------
        self.menu = tk.Menu(self, tearoff=0)
        self.menu.add_command(label="Description", command=self.show_description_popup)
        self.menu.add_separator()
        self.menu.add_command(label="Supprimer", command=self.delete_selected_row)

        # --------- Double clic "saisie magellan" -------------------
        self.tree.bind("<Button-3>", self.on_right_click)

        # --------- Charger les données -----------------------------
        self.load_data()

    # --- Méthode pour le filtrage années et mois -------------------
    def refresh_treeview(self):
        self.display_entries(self.all_entries)

    # ------------ Méthode pour la recherche d'organes --------------
    def search_organ(self):
        query = self.search_var.get().strip().lower()

        if not query:
            self.refresh()
            return

        filtered = [
            e for e in self.all_entries
            if query in e.get("materiel", "").lower()
            or query in e.get("sous_materiel", "").lower()
        ]

        self.display_entries(filtered)

    # ------------- Méthode afficher menu clic droit ----------------
    def on_right_click(self, event):
        item = self.tree.identify_row(event.y)

        if not item:
            return

        # Sélectionner la ligne sous la souris
        self.tree.selection_set(item)

        # Ne pas permettre suppression des séparateurs
        if "separator" in self.tree.item(item, "tags"):
            return

        # Afficher le menu
        self.menu.tk_popup(event.x_root, event.y_root)

    # --------- Supprimer la ligne via clic droit -------------------
    def delete_selected_row(self):
        selected = self.tree.selection()
        if not selected:
            return

        item = selected[0]
        values = self.tree.item(item, "values")

        jour, heure_debut, heure_fin = values[0], values[1], values[2]

        if not messagebox.askyesno(
            "Confirmation",
            "Voulez-vous vraiment supprimer cette intervention ?"
        ):
            return

        self.delete_entry_from_file(jour, heure_debut, heure_fin)
        self.refresh()

    # --------------- Supprimer definitivement la ligne -------------
    def delete_entry_from_file(self, jour, heure_debut, heure_fin):
        if not DATA_FILE.exists():
            return

        new_lines = []

        with open(DATA_FILE, "r", encoding="utf-8") as f:
            for line in f:
                if not line.strip():
                    continue
                try:
                    data = json.loads(line)
                except:
                    continue

                # On garde toutes les lignes SAUF celle à supprimer
                if (
                    data.get("jour") == jour and
                    data.get("heure_debut") == heure_debut and
                    data.get("heure_fin") == heure_fin
                ):
                    continue

                new_lines.append(json.dumps(data, ensure_ascii=False))

        with open(DATA_FILE, "w", encoding="utf-8") as f:
            for line in new_lines:
                f.write(line + "\n")

    # ----- Méthode pour afficher le popup description
    def show_description_popup(self):
        selected = self.tree.selection()
        if not selected:
            return

        item = selected[0]

        # Ne pas ouvrir sur une ligne séparatrice
        if "separator" in self.tree.item(item, "tags"):
            return

        description = self.descriptions.get(item, "").strip()
        if not description:
            description = "Aucune description."

        popup = ctk.CTkToplevel(self)
        popup.title("Description de l'intervention")
        popup.iconbitmap("ui/assets/train.ico")
        popup.geometry("500x320")
        popup.grab_set()

        popup.update_idletasks()
        self.center_popup(popup)

        # ----- Contenu -----
        frame = ctk.CTkFrame(popup)
        frame.pack(fill="both", expand=True, padx=15, pady=15)

        ctk.CTkLabel(
            frame,
            text="Description",
            font=("Roboto", 18)
        ).pack(pady=(5, 10))

        textbox = ctk.CTkTextbox(
            frame,
            wrap="word"
        )
        textbox.pack(fill="both", padx=10, expand=True)
        textbox.insert("1.0", description)
        textbox.configure(state="disabled")  # lecture seule

        ctk.CTkButton(
            frame,
            text="Fermer",
            command=popup.destroy,
            width=120
        ).pack(pady=(10, 10))

    # ------------ Méthode utilitaire : centrer un popup ------------
    def center_popup(self, popup):
        popup.update_idletasks()

        # Récupérer la position et taille réelle de la fenêtre principale
        main_x = self.winfo_rootx()
        main_y = self.winfo_rooty()
        main_width = self.winfo_width()
        main_height = self.winfo_height()

        # Taille réelle du popup
        popup_width = popup.winfo_width()
        popup_height = popup.winfo_height()

        # Calcul parfait du centre
        x = main_x + (main_width // 2) - (popup_width // 2)
        y = main_y + (main_height // 2) - (popup_height // 2)
        x -= 220
        y -= 50

        popup.geometry(f"+{x}+{y}")

    # ---------------------------------------------------------------------
    # Charger et afficher les données
    # ---------------------------------------------------------------------
    def load_data(self):
        self.tree.delete(*self.tree.get_children())
        self.descriptions.clear()

        if not DATA_FILE.exists():
            return

        entries = []
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    try:
                        entries.append(json.loads(line))
                    except json.JSONDecodeError:
                        continue
        self.all_entries = entries.copy()

        # ---- Mise à jour des années disponibles dans le dropdown ----
        years = sorted({
            e["jour"].split("-")[0]
            for e in self.all_entries
            if "jour" in e
        })

        self.year_combo.configure(values=years)

        # Sécurité : si l'année sélectionnée n'existe plus
        if years and self.selected_year.get() not in years:
            self.selected_year.set(years[-1])

        # 1) Calcul cumul journalier
        cumul_par_jour = {}
        for e in entries:
            jour = e.get("jour")
            temps = e.get("temps_total", 0)
            cumul_par_jour[jour] = cumul_par_jour.get(jour, 0) + temps

        # 2) Tri
        entries.sort(key=lambda x: (x["jour"], x["heure_debut"]))

        last_jour = None  # Pour détecter le changement de journée

        self.tree.tag_configure("separator", background="black")  # ligne noire
        
        # 3) Insert
        row_index = 0

        for entry in entries:
            jour = entry.get("jour", "")

            if last_jour is not None and jour != last_jour:
                self.tree.insert(
                    "",
                    "end",
                    iid=f"sep_{row_index}",
                    values=("", "", "", "", "", "", "", ""),
                    tags=("separator",)
                )
                row_index += 1

            last_jour = jour

            total = entry.get("temps_total", 0)
            temps_format = f"{total//60}h{total%60:02d}"

            # HV basé sur le cumul réel de la journée, en ignorant les entrées réinitialisées
            total_jour = sum(
                e.get("temps_total", 0)
                for e in self.all_entries
                if e.get("jour") == entry.get("jour") and not e.get("hv_reset", False)
            )
            hv_minutes = max(0, total_jour - 454)
            hv_format = f"+{hv_minutes//60}h{hv_minutes%60:02d}"

            iid = f"row_{row_index}"

            self.tree.insert(
                "",
                "end",
                iid=iid,
                values=(
                    entry.get("jour", ""),
                    entry.get("heure_debut", ""),
                    entry.get("heure_fin", ""),
                    temps_format,
                    hv_format,
                    entry.get("materiel", ""),
                    entry.get("sous_materiel", ""),
                    "Oui" if entry.get("saisie_magellan", False) else "Non"
                ),
                tags=("magellan_oui" if entry.get("saisie_magellan", False) else "magellan_non",)
            )

            self.descriptions[iid] = entry.get("description", "").strip()
            row_index += 1

        self.all_entries = entries.copy()  # stocker toutes les entrées
        self.display_entries(self.all_entries)

    # ----- Méthode pour la fonction de recherche -------
    def display_entries(self, entries):
        self.tree.delete(*self.tree.get_children())
        self.descriptions.clear()

        if not entries:
            return

        year = self.selected_year.get()
        month = self.selected_month.get()

        entries = [
            e for e in entries
            if e.get("jour", "").startswith(f"{year}-{month}")
        ]

        # --- Calcul cumul journalier sur toutes les entrées (pour HV correct)
        cumul_par_jour = {}
        for e in self.all_entries:  # toutes les entrées, pas juste filtrées
            jour = e.get("jour")
            temps = e.get("temps_total", 0)
            cumul_par_jour[jour] = cumul_par_jour.get(jour, 0) + temps

        # --- Tri par date et heure
        entries.sort(key=lambda x: (x["jour"], x["heure_debut"]))

        last_jour = None
        self.tree.tag_configure("separator", background="black")
        row_index = 0

        for entry in entries:
            jour = entry.get("jour", "")

            # Ligne séparatrice si changement de jour
            if last_jour is not None and jour != last_jour:
                self.tree.insert(
                    "",
                    "end",
                    iid=f"sep_{row_index}",
                    values=("", "", "", "", "", "", "", ""),
                    tags=("separator",)
                )
                row_index += 1

            last_jour = jour

            total = entry.get("temps_total", 0)
            temps_format = f"{total//60}h{total%60:02d}"

            # HV basé sur le cumul réel de la journée, en ignorant les entrées réinitialisées
            total_jour = sum(
                e.get("temps_total", 0)
                for e in self.all_entries
                if e.get("jour") == entry.get("jour") and not e.get("hv_reset", False)
            )
            hv_minutes = max(0, total_jour - 454)  # 454 = 7h34min
            hv_format = f"+{hv_minutes//60}h{hv_minutes%60:02d}"

            iid = f"row_{row_index}"
            self.tree.insert(
                "",
                "end",
                iid=iid,
                values=(
                    entry.get("jour", ""),
                    entry.get("heure_debut", ""),
                    entry.get("heure_fin", ""),
                    temps_format,
                    hv_format,
                    entry.get("materiel", ""),
                    entry.get("sous_materiel", ""),
                    "Oui" if entry.get("saisie_magellan", False) else "Non"
                ),
                tags=("magellan_oui" if entry.get("saisie_magellan", False) else "magellan_non",)
            )

            # --- ATTENTION ici : toujours remplir self.descriptions avec le bon iid
            self.descriptions[iid] = entry.get("description", "").strip()
            row_index += 1
        
    # ---------------------------------------------------------------------
    # Double clic → changer "saisie_magellan"
    # ---------------------------------------------------------------------
    def on_double_click(self, event):
        item = self.tree.selection()
        if not item:
            return

        col = self.tree.identify_column(event.x)
        col_index = int(col.replace("#","")) - 1

        # colonne saisie_magellan = dernière colonne (index 7)
        if col_index != 7:
            return

        values = self.tree.item(item, "values")
        jour, h_debut, h_fin = values[0], values[1], values[2]
        current_value = values[7]

        popup = ctk.CTkToplevel(self)
        popup.title("Modifier la saisie Magellan")
        popup.iconbitmap("ui/assets/train.ico")
        popup.geometry("320x150")
        popup.grab_set()

        popup.update_idletasks()
        self.center_popup(popup)

        ctk.CTkLabel(popup, text="Saisie dans Magellan ?").pack(pady=10)

        var = ctk.BooleanVar(value=(current_value == "Oui"))
        switch = ctk.CTkSwitch(popup, text="", variable=var)
        switch.pack(pady=10)

        def apply():
            self.update_magellan_status(jour, h_debut, h_fin, var.get())
            popup.destroy()

        ctk.CTkButton(popup, text="Valider", command=apply).pack(pady=10)

    # ---------------------------------------------------------------------
    # Modifier la valeur dans data.jsonl
    # ---------------------------------------------------------------------
    def update_magellan_status(self, jour, heure_debut, heure_fin, new_value):
        if not DATA_FILE.exists():
            return

        new_lines = []

        with open(DATA_FILE, "r", encoding="utf-8") as f:
            for line in f:
                if not line.strip():
                    continue
                try:
                    data = json.loads(line)
                except:
                    continue

                if (data.get("jour") == jour and 
                    data.get("heure_debut") == heure_debut and 
                    data.get("heure_fin") == heure_fin):
                    data["saisie_magellan"] = new_value

                new_lines.append(json.dumps(data, ensure_ascii=False))

        # Réécriture fichier
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            for line in new_lines:
                f.write(line + "\n")

        self.refresh()

    # Exporter le tableau de l'historique sur une feuille EXCEL
    def export_excel(self):
        if not DATA_FILE.exists():
            messagebox.showinfo("Export Excel", "Aucune donnée à exporter.")
            return

        selected_year = self.selected_year.get()

        # -------- Lecture JSONL --------
        entries = []
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            for line in f:
                if line.strip():
                    entries.append(json.loads(line))

        # -------- Filtre par année --------
        entries_year = [
            e for e in entries
            if e.get("jour", "").startswith(selected_year)
        ]

        if not entries_year:
            messagebox.showinfo("Export Excel", "Aucune donnée pour cette année.")
            return

        # -------- Regroupement par mois --------
        rows_by_month = defaultdict(list)

        # -------- Calcul HV cumulées comme page accueil --------
        MAX_JOURNEE = 454
        MAX_HV_PAR_JOUR = 454

        hv_par_jour = {}

        entries_by_day = defaultdict(list)

        for e in entries_year:
            jour = e.get("jour")
            if jour:
                entries_by_day[jour].append(e)

        for jour in sorted(entries_by_day.keys()):

            total_jour = sum(
                e.get("temps_total", 0)
                for e in entries_by_day[jour]
            )

            hv_jour = 0

            if total_jour > MAX_JOURNEE:
                excedent = total_jour - MAX_JOURNEE
                hv_jour = min(excedent, MAX_HV_PAR_JOUR)

            hv_par_jour[jour] = hv_jour

        # -------- Génération des lignes Excel --------
        for jour in sorted(entries_by_day.keys()):
            for e in entries_by_day[jour]:

                month_key = e["jour"][:7]

                # Temps total formaté
                temps_total = e.get("temps_total", 0)
                temps_format = f"{temps_total//60}h{temps_total%60:02d}"

                # HV formatées
                hv_minutes = hv_par_jour.get(jour, 0)
                hv_format = f"+{hv_minutes//60}h{hv_minutes%60:02d}"

                rows_by_month[month_key].append([
                    format_date_fr(jour),
                    e.get("heure_debut"),
                    e.get("heure_fin"),
                    temps_format,
                    hv_format,
                    e.get("materiel", ""),
                    e.get("sous_materiel", ""),
                    "oui" if e.get("saisie_magellan") else "non"
                ])

        #file_str = date.today().strftime("%Y-%m-%d")
        default_filename = f"Suivi_journees_activites_travail_RATP-{selected_year}.xlsx"

        filepath = asksaveasfilename(
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Fichier Excel", "*.xlsx")]
        )
        if not filepath:
            return

        wb = Workbook()
        wb.remove(wb.active)

        headers = [
            "Jour", "Heure début", "Heure fin", "Temps total",
            "HV", "Organe", "Sous organe", "Saisie Magellan"
        ]

        for month_key, month_rows in sorted(rows_by_month.items()):
            sheet_name = format_month_fr(month_key)
            ws = wb.create_sheet(title=sheet_name)

            # -------- Couleur onglet --------
            month_num = month_key.split("-")[1]
            if month_num in MONTH_COLORS:
                ws.sheet_properties.tabColor = MONTH_COLORS[month_num]

            # -------- Titre --------
            ws.merge_cells("A1:H2")
            title_cell = ws["A1"]
            title_cell.value = "Historique des interventions RATP"
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.fill = PatternFill(fill_type="solid", fgColor="FFFF00")

            # -------- En-têtes --------
            header_row = 5
            thin = Side(border_style="thin", color="000000")

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(header_row, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # -------- Contenu --------
            current_row = header_row + 1
            last_day = None

            for row in month_rows:
                if last_day and row[0] != last_day:
                    for col in range(1, 9):
                        ws.cell(row=current_row, column=col).fill = PatternFill(
                            start_color="EEEEEE", fill_type="solid"
                        )
                    current_row += 1

                last_day = row[0]

                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                    if col_num == 8 and str(value).lower() == "non":
                        cell.fill = PatternFill(start_color="FFD966", fill_type="solid")

                current_row += 1

            # -------- Largeur colonnes --------
            for i in range(1, 9):
                col_letter = ws.cell(row=header_row, column=i).column_letter
                ws.column_dimensions[col_letter].width = 18

        wb.save(filepath)
        messagebox.showinfo("Export Excel", "Export Excel annuel réussi.")

    # Exporter le tableau en fichier PDF
    def export_pdf(self):
        if not DATA_FILE.exists():
            messagebox.showinfo("Export PDF", "Aucune donnée à exporter.")
            return

        selected_year = self.selected_year.get()

        # ---------- Lecture JSONL ----------
        entries = []
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            for line in f:
                if line.strip():
                    entries.append(json.loads(line))

        # ---------- Filtre par année ----------
        entries_year = [
            e for e in entries
            if e.get("jour", "").startswith(selected_year)
        ]

        if not entries_year:
            messagebox.showinfo("Export PDF", "Aucune donnée pour cette année.")
            return

        MAX_JOURNEE = 454
        MAX_HV_PAR_JOUR = 454

        hv_par_jour = {}
        entries_by_day = defaultdict(list)

        for e in entries_year:
            jour = e.get("jour")
            if jour:
                entries_by_day[jour].append(e)

        for jour in sorted(entries_by_day.keys()):

            total_jour = sum(
                e.get("temps_total", 0)
                for e in entries_by_day[jour]
            )

            hv_jour = 0

            if total_jour > MAX_JOURNEE:
                excedent = total_jour - MAX_JOURNEE
                hv_jour = min(excedent, MAX_HV_PAR_JOUR)

            hv_par_jour[jour] = hv_jour

        # ---------- Regroupement par mois ----------
        rows_by_month = defaultdict(list)

        for e in entries_year:

            jour = e.get("jour")

            # format temps total
            temps_total = e.get("temps_total", 0)
            temps_format = f"{temps_total//60}h{temps_total%60:02d}"

            # HV
            hv_minutes = hv_par_jour.get(jour, 0)
            hv_format = f"+{hv_minutes//60}h{hv_minutes%60:02d}" if hv_minutes > 0 else ""

            magellan = e.get("saisie_magellan", False)
            magellan_txt = "oui" if magellan in [True, "True", "true", "oui"] else "non"

            rows_by_month[jour[:7]].append([
                format_date_fr(jour),
                e.get("heure_debut"),
                e.get("heure_fin"),
                temps_format,
                hv_format,
                e.get("materiel", ""),
                e.get("sous_materiel", ""),
                magellan_txt
            ])

        default_filename = f"Suivi_journees_activites_travail_RATP-{selected_year}.pdf"

        filepath = asksaveasfilename(
            initialfile=default_filename,
            defaultextension=".pdf",
            filetypes=[("Fichier PDF", "*.pdf")]
        )
        if not filepath:
            return

        # ---------- Mise en page ----------
        width, height = A4
        margin = 20
        row_height = 18
        col_widths = [60, 45, 45, 45, 45, 95, 95, 50]
        table_width = sum(col_widths)
        x_table = (width - table_width) / 2

        c = canvas.Canvas(filepath, pagesize=A4)

        header = ["Jour", "Début", "Fin", "Total", "HV", "Organe", "Sous organe", "Magellan"]

        page_number = 1

        def draw_footer():
            now = datetime.now().strftime("%d/%m/%Y à %H:%M")
            c.setFont("Times-Italic", 9)
            c.setFillColor(HexColor("#666666"))
            c.drawCentredString(
                width / 2, 20,
                f"Généré le {now} - Page {page_number}"
            )
        
        def draw_header_block(month):
            box_y = height - 70
            c.setFillColor(HexColor("#1E5CC4"))
            c.roundRect(margin, box_y, width - 2 * margin, 45, 8, fill=1)

            try:
                logo = ImageReader("ui/assets/logo_ratp.png")
                c.drawImage(logo, margin + 10, box_y + 2, width=40, height=40, mask="auto")
            except:
                pass

            c.setFont("Helvetica-Bold", 22)
            c.setFillColor(HexColor("#FFFFFF"))
            c.drawCentredString(width / 2, box_y + 13, f"Historique {month}")

            y = box_y - 40

            # En-têtes tableau
            c.setFillColor(HexColor("#1E5CC4"))
            c.rect(x_table, y - row_height, table_width, row_height, fill=1)
            c.setFont("Times-Bold", 10)
            c.setFillColor(HexColor("#FFFFFF"))

            col_x = x_table
            for i, h in enumerate(header):
                c.drawString(col_x + 4, y - row_height + 5, h)
                col_x += col_widths[i]

            return y - row_height

        # ---------- Génération ----------
        for month, rows in sorted(rows_by_month.items()):
            year, month_num = month.split("-")
            month_label = f"{MONTHS_FR.get(month_num, month_num)} {year}"
            y = draw_header_block(month_label)
            c.setFont("Times-Roman", 9)
            last_day = None

            for idx, row in enumerate(rows):
                # Pagination dynamique (CORRECTION CLÉ)
                if y < 60:
                    draw_footer()
                    c.showPage()
                    page_number += 1
                    y = draw_header_block(month_label)
                    c.setFont("Times-Roman", 9)
                    last_day = None

                # Séparateur journée
                if last_day and row[0] != last_day:
                    c.setFillColor(HexColor("#BFBFBF"))
                    c.rect(x_table, y - row_height, table_width, row_height, fill=1)
                    y -= row_height

                last_day = row[0]

                # Couleur ligne
                if row[7] == "non":
                    c.setFillColor(HexColor("#FFD966"))
                elif idx % 2 == 0:
                    c.setFillColor(HexColor("#F2F2F2"))
                else:
                    c.setFillColor(HexColor("#FFFFFF"))

                c.rect(x_table, y - row_height, table_width, row_height, fill=1)

                c.setFillColor(HexColor("#000000"))
                col_x = x_table
                for i, value in enumerate(row):
                    c.drawString(col_x + 4, y - row_height + 5, str(value))
                    col_x += col_widths[i]

                y -= row_height

            draw_footer()
            c.showPage()
            page_number += 1

        c.save()
        messagebox.showinfo("Export PDF", "Export PDF annuel réussi.")

    # ---------------- Rafraichissement -----------------------------------
    def refresh(self):
        self.load_data()

    # ---------------- Effacer les données confirmation -------------------
    def clear_data(self):
        if messagebox.askyesno("Confirmation", "Voulez-vous vraiment supprimer toutes les données ?"):
            with open(DATA_FILE, "w", encoding="utf-8") as f:
                f.write("")
            self.tree.delete(*self.tree.get_children())
    
    # Traduction du language
    def refresh_language(self):
        self.lang_util = PageLang(self.app)
        self.header_label.configure(text=self.lang_util.t("historique_titre"))
        self.search_label.configure(text=self.lang_util.t("rechercher_organe"))
        self.btn_search.configure(text=self.lang_util.t("rechercher"))
        self.btn_reset.configure(text=self.lang_util.t("reinitialiser"))
        self.btn_deleteall.configure(text=self.lang_util.t("tout_supprim"))
        self.btn_refresh.configure(text=self.lang_util.t("rafraichir"))
        self.btn_excel.configure(text=self.lang_util.t("excel"))
        self.btn_pdf.configure(text=self.lang_util.t("pdf"))

        # Headers du Treeview
        for col, key in self.column_headers.items():
            self.tree.heading(col, text=self.lang_util.t(key))
